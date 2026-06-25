#!/usr/bin/env python3
"""
Generate Report Script — CryptoSentinel AI Trading Agent
Reads trade logs from backtest_log.csv and outputs a styled, professional
Excel report (report.xlsx) using openpyxl.
"""

import os
import csv
import sys
import argparse

# Configure standard output to use UTF-8 if possible
try:
    if sys.platform.startswith('win'):
        import sys
        sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def create_csv_fallback(csv_path, excel_path):
    print("[WARNING] openpyxl is not installed. Generating a clean Markdown/CSV text summary instead.")
    if not os.path.exists(csv_path):
        print(f"[ERROR] LOG FILE NOT FOUND: {csv_path}")
        return

    trades = []
    with open(csv_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            trades.append(row)

    if not trades:
        print("[ERROR] No trades to report.")
        return

    # Calculate basic stats
    total_trades = len(trades)
    winning_trades = sum(1 for t in trades if float(t["pnl_usdt"] or 0) > 0)
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    total_pnl = sum(float(t["pnl_usdt"] or 0) for t in trades)
    
    summary_path = excel_path.replace(".xlsx", "_summary.txt")
    with open(summary_path, "w") as f:
        f.write("==================================================\n")
        f.write("         CRYPTOSENTINEL BACKTEST SUMMARY\n")
        f.write("==================================================\n")
        f.write(f"Total Trades:      {total_trades}\n")
        f.write(f"Winning Trades:    {winning_trades}\n")
        f.write(f"Win Rate:          {win_rate:.1f}%\n")
        f.write(f"Total Net PnL:     ${total_pnl:+.2f} USDT\n")
        f.write("==================================================\n")
        f.write("\nDetailed trades have been logged in backtest_log.csv\n")
    
    print(f"[OK] Summary written to {summary_path}")

def generate_excel_report(csv_path, excel_path):
    if not os.path.exists(csv_path):
        print(f"[ERROR] LOG FILE NOT FOUND: {csv_path}")
        return

    # Load trades
    trades = []
    with open(csv_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            trades.append(row)

    if not trades:
        print("[ERROR] No trades to report.")
        return

    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Performance Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Stylings
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="3B1E6D", end_color="3B1E6D", fill_type="solid") # Deep purple
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B3F8F", end_color="5B3F8F", fill_type="solid") # Secondary purple
    accent_fill = PatternFill(start_color="F2EFFF", end_color="F2EFFF", fill_type="solid") # Very light purple
    
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    normal_font = Font(name="Segoe UI", size=11)
    green_font = Font(name="Segoe UI", size=11, bold=True, color="008F00")
    red_font = Font(name="Segoe UI", size=11, bold=True, color="9F0000")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Title Banner
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "CryptoSentinel Trading Performance Report"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = title_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40

    # Summary Stats computation
    total_trades = len(trades)
    winning_trades = sum(1 for t in trades if float(t["pnl_usdt"] or 0) > 0)
    losing_trades = total_trades - winning_trades
    win_rate = winning_trades / total_trades if total_trades > 0 else 0
    total_pnl = sum(float(t["pnl_usdt"] or 0) for t in trades)
    
    # Calculate profit factor
    gross_profits = sum(float(t["pnl_usdt"] or 0) for t in trades if float(t["pnl_usdt"] or 0) > 0)
    gross_losses = abs(sum(float(t["pnl_usdt"] or 0) for t in trades if float(t["pnl_usdt"] or 0) < 0))
    profit_factor = gross_profits / gross_losses if gross_losses > 0 else float("inf")

    stats = [
        ("Key Metric", "Value", "Notes"),
        ("Total Trades", total_trades, "Total executed paper trades"),
        ("Winning Trades", winning_trades, "Trades with positive PnL"),
        ("Losing Trades", losing_trades, "Trades with negative/neutral PnL"),
        ("Win Rate", f"{win_rate * 100:.1f}%", "Percentage of winning trades"),
        ("Total Net PnL", f"${total_pnl:+.2f} USDT", "Total cumulative profit/loss"),
        ("Profit Factor", f"{profit_factor:.2f}" if profit_factor != float("inf") else "N/A", "Gross Profits / Gross Losses")
    ]

    for idx, (metric, val, note) in enumerate(stats, start=3):
        ws_summary.cell(row=idx, column=1, value=metric)
        ws_summary.cell(row=idx, column=2, value=val)
        ws_summary.cell(row=idx, column=3, value=note)
        
        # Apply style
        for col in (1, 2, 3):
            cell = ws_summary.cell(row=idx, column=col)
            cell.font = normal_font if idx > 3 else header_font
            cell.border = thin_border
            if idx == 3:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                if col == 1:
                    cell.font = bold_font
                    cell.fill = accent_fill
                elif col == 2:
                    cell.alignment = Alignment(horizontal="right")
                    if "Total Net PnL" in metric:
                        cell.font = green_font if total_pnl >= 0 else red_font

    # Column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 35

    # 2. Trade History Sheet
    ws_history = wb.create_sheet(title="Trade History")
    ws_history.views.sheetView[0].showGridLines = True
    
    headers = [
        "Timestamp", "Asset", "Action", "Amount (USDT)", "Entry Price",
        "Stop Loss", "Take Profit", "Signal Score", "Confidence",
        "Exit Price", "PnL (USDT)", "PnL (%)", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_history.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_history.row_dimensions[1].height = 25

    # Fill History Data
    for row_idx, trade in enumerate(trades, start=2):
        row_data = [
            trade.get("timestamp", ""),
            trade.get("asset", ""),
            trade.get("action", ""),
            float(trade.get("amount_usdt", 0) or 0),
            float(trade.get("entry_price", 0) or 0),
            float(trade.get("stop_loss", 0) or 0),
            float(trade.get("take_profit", 0) or 0),
            float(trade.get("signal_score", 0) or 0) if trade.get("signal_score") != "N/A" else "N/A",
            trade.get("confidence", ""),
            float(trade.get("exit_price", 0) or 0) if trade.get("exit_price") else "",
            float(trade.get("pnl_usdt", 0) or 0) if trade.get("pnl_usdt") else "",
            float(trade.get("pnl_pct", 0) or 0) if trade.get("pnl_pct") else "",
            trade.get("status", "")
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_history.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formattings
            if col_idx in (4, 5, 6, 7, 10, 11):
                if val != "" and val != "N/A":
                    cell.number_format = "$#,##0.00" if col_idx != 8 else "0.0"
                    cell.alignment = Alignment(horizontal="right")
            elif col_idx == 12:
                if val != "":
                    cell.number_format = "+0.0%;-0.0%;0.0%"
                    cell.alignment = Alignment(horizontal="right")
            elif col_idx in (2, 3, 9, 13):
                cell.alignment = Alignment(horizontal="center")
                
            # Colored PnL
            if col_idx in (11, 12):
                if val != "":
                    cell.font = green_font if val >= 0 else red_font

    # Auto-fit columns
    for col in ws_history.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_history.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    print(f"[OK] Excel report generated successfully at {excel_path}")

def main():
    parser = argparse.ArgumentParser(description="Generate Excel performance report.")
    parser.add_argument("--input", default="backtest_log.csv", help="Input CSV log file path")
    parser.add_argument("--output", default="report.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    if OPENPYXL_AVAILABLE:
        generate_excel_report(args.input, args.output)
    else:
        create_csv_fallback(args.input, args.output)

if __name__ == "__main__":
    main()
